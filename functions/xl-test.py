import datetime
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1, 2, 3])

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("./test/sample.xlsx")